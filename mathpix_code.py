import os
import cv2
import json
import re
import matplotlib.pyplot as plt
import requests
import base64
from PIL import Image as ImagePIL
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import time
import shutil

url = 'https://api.mathpix.com/v3/text'
app_id = 'siby_sebastian_magicsw_com_secure'
app_key = '47252dbebd2a29ddd3ee701d4691d6bcb0aa94c6c2b869f0850e9470e301bda3'

# Request headers
headers = {
    'app_id': app_id,
    'app_key': app_key,
    'Content-type': 'application/json'
}

# List of allowed LaTeX symbols
latex_symbols = [
    "-","+","\\times", "\\pm", "\\mp", "\\div", "\\cdot", "\\neq", "\\geq", "\\leq", "\\theta", "\\lambda", "\\mu", "\\pi",
    "=", "<", ">", "_",
    "^", "|", "\\infty", "\\div", "\\sqrt", "\\pm", "\\mp", "\\cdot",
    "\\neq", "\\geq", "\\leq", "\\theta", "\\lambda", "\\mu", "\\pi",
    "\\alpha", "\\beta", "\\gamma", "\\phi", "\\Sigma", "\\Omega",
    "\\nabla", "\\int", "\\sum", "\\prod", "\\subset", "\\cup", "\\cap",
    "\\rightarrow", "\\leftarrow", "\\Rightarrow", "\\Leftarrow",
    "\\leftrightarrow", "\\Leftrightarrow", "= -", "\\frac", "\\emptyset","\\text","\\quad"
]

workbook = Workbook()
worksheet = workbook.active
# Define the starting row
current_row = 2  # Start from row 2 to allow space for headers


def resizeImg(folderPath,png_file):
    # Open the image
    imagepath = folderPath + "\\" + png_file
    print("imagepath: ",imagepath)
    print("pngFile: ",png_file)
    img = ImagePIL.open(imagepath)
    # Resize the image to a smaller dimension (e.g., 50% of the original size)
    img = img.resize((img.width // 2, img.height // 2))

    # Save the resized image
    resizeImagePath = folderPath + "\\" + "resize_" + png_file
    img.save(resizeImagePath)
    return resizeImagePath

def is_number(value):
    try:
        int(value)
        return True
    except ValueError:
        return False

def callMathpixAPI(file_path):
        with open(file_path, 'rb') as image_file:
            image_data = base64.b64encode(image_file.read()).decode('utf-8')
                    # Request body
        data = {
                'src': 'data:image/png;base64,' + image_data,
                "formats": ["text", "data", "html", "latex_styled","math"],
                "include_asciimath": True,
                 "include_line_data": True,
                 #"include_word_data": True,
                "include_latex": True,
                "include_mathml": True
            }


            # Send the POST request to the Mathpix API
        response = requests.post(url, json=data, headers=headers)
    #        result = {}
    #         with open("errordata.json", "r") as file:
    #             resultFileStr = file.read()
    #             result = json.loads(resultFileStr)

        result = response.json()
        print('JSON:',result)
        return result

# find_start_end to take a worksheet argument and populate the column C
def find_start_end(text, substring, worksheet):
    # print("text",text)
    # print("substring", substring)
    start = text.find(substring)
    if start != -1:
        end = start + len(substring) - 1
        position_str = f"Start: {start}, End: {end}"
        print(f"'{substring}' starts at position {start} and ends at position {end}.")
        # Add the position string to the third column (column C)
        worksheet[f'C{current_row}'] = position_str
    else:
        print(f"'{substring}' not found in the text.")

    
    

def generateLatexCode(result,png_file,worksheet,save_path):
    count = 0
    line_number = 0
    global current_row
    
    errorObj = result.get('error')
    
    if errorObj is not None and 'Image too large' == errorObj:
        print('error image is large')
        raise ValueError("Image too large")
        
    for word_entry in result['line_data']:
        line_number = line_number + 1
        cnt_data = word_entry['cnt']
        if 'text' in word_entry:
            cnt_word = word_entry['text']
            contains_math = any(symbol in cnt_word for symbol in latex_symbols)
            contains_numbers = any(is_number(char) for char in cnt_word)
            print("Text",cnt_word)
            # print("LineNumber",line)
            if contains_math and contains_numbers:
                #print("line_number",line_number)
                #Add the line number to the second column (column B)
                latex_line = word_entry.get('text').strip()
                latex_line = latex_line.replace('stackrel', 'overset')
                # if "text" in latex_line:
                #     # latex_line = "a^{\frac{1}{2}}=\overset{?}{?} \cdot \sqrt{a}"
                #     continue
                # if 'entonces' in latex_line:
                #     #latex_line = "a^{\frac{1}{2}}=\overset{?}{?} \cdot \sqrt{a}"
                #     continue
                # if 'boldsymbol' in latex_line:
                #     continue
                # if '\\right' in latex_line:
                #     continue
                # if '\\underline' in latex_line:
                #     continue
                # if '\\square' in latex_line:
                #     continue
                # if '\\longdiv' in latex_line:
                #     continue
                # if '\\begin' in latex_line:
                #     continue
                # if '\\triangle' in latex_line:
                #     continue
                # if '\sqrt{ }' in latex_line:
                    # continue
                #latex_line = latex_line.replace('text', ' ')

                # print('Text Identified :: '+ latex_line)
                # Remove \( and \)

                latex_line = latex_line.replace('\(', '').replace('\)', '')

                # Function to wrap English words in \text{}
                def wrap_english_words(match):
                    word = match.group(0)
                    return f'\\text{{{word}}}'

                # Regular expression pattern to match English words not starting with \
                pattern = r'(?<!\\)(?<!{)\b[a-zA-Z]+\b'

                # Use re.sub() to replace English words with \text{}
                latex_line= re.sub(pattern, wrap_english_words, latex_line)
                latex_line = latex_line.replace('} ', '}\ ')

                worksheet[f'B{current_row}'] = line_number
                print("line",line_number)
                latex_code = latex_line
                print("final latex code :: "+latex_code)
                print("Coordinates:",cnt_data)
                coordinates=[cnt_data[0][0],cnt_data[2][0],cnt_data[1][1],cnt_data[0][1]]
                worksheet[f'C{current_row}'] = str(coordinates)


                # Create a larger figure with adjusted margins
                fig, ax = plt.subplots(figsize=(3, 1))
                fig.subplots_adjust(left=0.5, right=0.9, top=0.5, bottom=0.1)

                # Add a LaTeX equation as an annotation
                ax.annotate(latex_code, xy=(0, 1), fontsize=12, ha='center' , va='center')

                # Turn off the axis
                ax.axis('off')

                # Show the plot
                count = count+1
                timestamp = int(time.time())  # Get current timestamp
                filename = f"eq_{timestamp}_{count}.png"

                if not os.path.exists(save_path):
                    os.makedirs(save_path)

                filename = os.path.join(save_path, filename)

                # print("filename",filename)
                
                plt.savefig(filename)
                # plt.show()
                worksheet[f'A{current_row}'] = str(png_file)
                # print("page:",png_file)
                
                # Add the starting and ending character position to the third column (column C)
                # find_start_end(latex_line, latex_code, worksheet)
                
                # Add the image to the E column (column E)
                img = Image(filename)
                img.width = 150 # Set the image width (adjust as needed)
                img.height = 60  # Set the image height (adjust as needed)
                worksheet.add_image(img, f'D{current_row}')

                # Add the text to the sixth column (column F)
                worksheet[f'E{current_row}'] = latex_code
                # Add image number to the seventh column (column G)
                worksheet[f'F{current_row}'] = png_file.split("_")[-1].split(".")[0]
                # Increment the row for the next iteration
                current_row = current_row + 1
                print("Current:",current_row)
                print("\n")
                        
                        

def mathpixAPI(folder_path,output_path):

    save_path=os.path.join(output_path,"mathpix_output")
    if os.path.exists(save_path) and os.path.isdir(save_path):
        shutil.rmtree(save_path)
        
    if not os.path.exists(save_path):
        os.makedirs(save_path)

    # Set column widths (adjust as needed)
    worksheet.column_dimensions[get_column_letter(1)].width = 20
    worksheet.column_dimensions[get_column_letter(2)].width = 40

    # Create headers in the first row
    worksheet[f'A1'] = 'Page_No'
    worksheet[f'B1'] = 'Line_Number'
    # worksheet[f'C1'] = "Starting_Character&Ending_Character_Position"
    worksheet[f'C1']= 'Coordinates'
    worksheet[f'D1'] = 'Image'
    worksheet[f'E1'] = 'Latex'
    worksheet[f'F1'] = 'Slide_No'

    # Check if the folder exists
    if os.path.exists(folder_path) and os.path.isdir(folder_path):
        # List all files in the folder
        files = os.listdir(folder_path)

        # Filter for PNG files
        png_files = [file for file in files if file.endswith('.png')]
        sorted_image_list = sorted(png_files, key=lambda x: int(x.split("_")[-1].split(".")[0]))
        print(sorted_image_list)

        # Iterate over the PNG files
        for png_file in sorted_image_list:
            # Create the full file path
            file_path = os.path.join(folder_path, png_file)

            # Process the PNG file (e.g., display, analyze, or manipulate)
            print(f"Processing {file_path}")
            
            result = callMathpixAPI(file_path)
            
            try:
                generateLatexCode(result,png_file,worksheet,save_path)
            except ValueError as e:
                print(f"Error {e}")
                resized_img = resizeImg(folder_path,png_file)
                result =  callMathpixAPI(resized_img)
                generateLatexCode(result,png_file,worksheet,save_path)

    else:
        print(f"The folder '{folder_path}' does not exist.")
    
    excel_path=os.path.join(save_path,"Result.xlsx")
    
    workbook.save(excel_path)

if __name__ == '__main__':
    folder_path = r"D:\last_try"
    output_path = r"D:\last_try"
    mathpixAPI(folder_path,output_path)